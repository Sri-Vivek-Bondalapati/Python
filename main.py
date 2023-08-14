'''import csv
import openpyxl
import sys

csv.field_size_limit(sys.maxsize)
input_file = '/Users/srivivek/Study/Python/realData.txt'
output_file = 'new.csv'

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
val = open("/Users/srivivek/Study/Python/realData.txt",mode='rb')
data = val.read()
daladec =data.decode('latin-1')
ader = csv.reader(val, delimiter='')
print(ader)
for row in ader:
    ws.append(row)
wb.save(output_file)
#print(daladec)
'''
import csv
import openpyxl

input_file = '/Users/srivivek/Study/Python/realData.txt'
output_file = 'test.xlsx'

wb = openpyxl.Workbook()
sheet = wb.create_sheet(title='Unique')
ws = wb.worksheets[0]

with open(input_file, 'r',encoding='latin-1') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)
wb.save(output_file)    
